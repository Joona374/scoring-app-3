from fastapi import APIRouter, Response, Depends, Query
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import load_workbook
from collections import defaultdict
import copy

from utils import get_current_user_id
from db.db_manager import get_db_session
from db.models import TeamStatsTag, User, Game, PlayerStatsTag, ShotResult, ShotResultTypes, ShotAreaTypes, ShotTypeTypes, PlayerStatsTagOnIce, PlayerStatsTagParticipating

router = APIRouter(
    prefix="/excel",
    tags=["excel"],
    responses={404: {"description": "Not found"}},
)

# TODO: Remove this?
@router.get("/download-test")
async def download_excel():
    workbook = load_workbook("mock_excel.xlsx")
    player_template = workbook.active

    names = ["Pekka", "Tero", "Pasi", "Harri", "Kari"]

    for i in range(5):
        new_sheet = workbook.copy_worksheet(player_template)
        new_sheet.title = names[i]
        new_sheet["B1"] = names[i]
        new_sheet["D1"] = i * 2

    workbook.remove(workbook.worksheets[0])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stats.xlsx"}
    )


########################## FOR TEAM STATS ##########################
def get_result_column_shifter(scoring_chance: TeamStatsTag):
    if scoring_chance.play_result == "Maali +":
        return 0
    elif scoring_chance.play_result == "Maali -":
        return 3
    elif scoring_chance.play_result == "MP +":
        return 6
    elif scoring_chance.play_result == "MP -":
        return 9
    
def get_chance_row(scoring_chance: TeamStatsTag):
    try:
        CHANCE_ROW_MAPPING = {
            "rush_type1": {
                "Tasavoimainen": 10,
                "Ylivoimainen": 11,
                "Alivoimainen": 12,
                "Läpiajo": 13
                },

            "takeaway_type": {
                "HAPP/PAHP": 15,
                "KAPP/KAHP": 17,
                "PAPP/HAHP": 19,
                "Jatkopaine": 21,
            },

            "hahp_papp_type": {
                "Täyttö": 23,
                "Alapeli": 25,
                "Yläpeli": 27
            },

            "rebound_type": {
                "SHP": 29,
                "Riisto/Menetys": 29,
                "HAHP/PAPP": 29,
            },

            "faceoff_type": {
                "Hyökkäysalue": 31,
                "Keskialue": 31,
                "Puoulustusalue": 31,
            },

            "v5v5_other_type": {
                "IM": 33,
                "TM": 33,
                "4v4": 33,
            },

            "pp_faceoff_entry_type": {
                "Aloitus Vasen": 38,
                "Aloitus Oikea": 38,
                "Haku / vastaanotto": 38
            },

            "pp_shot_deflection_low_type1": {
                "Päädystä": 40,
                "Siiveltä": 41,
                "Keskeltä": 42,
                "Viivasta": 43
            },

            "pp_blueline_shot_type": {
                "Suora": 45,
                "Ohjuri": 45,
                "Rebound": 45
            },

            "pp_pressure_brokenplay_type": {
                "Paine": 47,
                "Riisto": 47,
                "Brokenplay": 47
            },

            "pp_other_type": {
                "Kuljetus": 49,
                "Punnerrus": 49,
                "Rebound": 49
            },

            "pp_5vs3_type": {
                "Suora": 51,
                "Ohjuri": 51,
                "Rebound": 51
            },

            "pp_av_yv_type": {
                "Läpiajo": 53,
                "YV": 53,
                "TV": 53
            },

            "v3vs3_type": {
                "Aloitus": 58,
                "Hallinta": 58,
                "Riisto / Menetys": 58
            },

            "ps_type": {
                "Laukaus": 60,
                "Harhautus": 60,
                "Muu": 60
            },
            }
        
        for row_name in CHANCE_ROW_MAPPING.keys():
            value = getattr(scoring_chance, row_name, None)
            if value:
                return CHANCE_ROW_MAPPING[row_name][value]
    except Exception as e:
        print(f"Error in get_chance_row with tag:\n{scoring_chance}\nError:{e}")

def get_chance_column(scoring_chance: TeamStatsTag):
    final_columns = [
        "rush_type2", "takeaway_happ_pahp_type", "takeaway_kapp_kahp_type", "takeaway_papp_hahp_type", 
        "takeaway_jatkopaine_type", "hahp_papp_taytto_type", "hahp_papp_alapeli_type", 
        "hahp_papp_ylapeli_type", "rebound_type", "faceoff_type", "v5v5_other_type", 
        "pp_faceoff_entry_type", "pp_shot_deflection_low_type2", "pp_blueline_shot_type", 
        "pp_pressure_brokenplay_type", "pp_other_type", "pp_5vs3_type", "pp_av_yv_type", "v3vs3_type", "ps_type"]
    
    g_columns = ["PAHP", "1. Paine", "Syöttö", "Pohja", "Murtautuminen", "Syöttö sisään", "Suora", "SHP", "Hyökkäysalue", 
                 "IM", "Aloitus Vasen", "Kesk. Pääty.", "Paine", "Kuljetus YV", "Läpiajo", "Aloitus", "Laukaus"]
    h_columns = ["KAHP", "2. Paine", "Kuljetus", "Half Board", "Syöttö 3", "Murtautuminen sisään", "Ohjaus", "Riisto/Menetys", 
                 "Keskialue", "TM", "Aloitus Oikea", "Vasen Siipi", "Ohjuri", "Riisto", "Punnerrus", "Ohjuri", "YV", "Hallinta", "Harhautus"]
    i_columns = ["Kääntö", "3. / Puolustajan Paine", "Muu", "Viiva", "Syöttö 4/5", "HAHP/PAPP", "Puolustusalue", "4v4", 
                 "Haku / vastaanotto", "Oikea siipi", "Rebound", "Brokenplay", "TV", "Riisto / Menetys"]

    for column in final_columns:
        value = getattr(scoring_chance, column, None)
        if value:
            if value in g_columns:
                return "G"
            elif value in h_columns:
                return "H"
            elif value in i_columns:
                return "I"
            else:
                print(scoring_chance)
                print(f"This is probelm: {value}")
                raise ValueError("COLUMN NOT FOUND ANYWHERE SOS! :D")

def calculate_numbers_for_cells(all_tags):
    cell_values = {}
    for tag in all_tags:
        row = get_chance_row(tag)
        col = get_chance_column(tag)
        col_shift = get_result_column_shifter(tag)
        shifted_col = chr(ord(col) + col_shift)

        if f"{shifted_col}{row}" in cell_values:
            cell_values[f"{shifted_col}{row}"] += 1
        else:
            cell_values[f"{shifted_col}{row}"] = 1
    
    return cell_values

@router.get("/teamstats")
async def get_teamstats_excel(game_ids: str, db_session: Session = Depends(get_db_session), current_user_id: int = Depends(get_current_user_id)):

    user = db_session.query(User).filter(User.id == current_user_id).first()
    team = user.team

    teams_games = team.games
    if game_ids:
        filter_game_ids = [int(game_id) for game_id in game_ids.split(",")]
    else:
        filter_game_ids = [game.id for game in teams_games]
    

    # Get all the tag wit db query
    all_tags = db_session.query(TeamStatsTag).filter(TeamStatsTag.game.has(team=user.team), TeamStatsTag.game_id.in_(filter_game_ids)).all()

    # Iterate over all_tags to get tags for individual games
    # games_with_stats_dict = {game_id: list[TeamStatsTag]} 
    games_with_stats_dict = defaultdict(list)
    for tag in all_tags:
        games_with_stats_dict[tag.game_id].append(tag)

    # Load the excel file and template sheet
    workbook = load_workbook("excels/team_stats_template.xlsx")
    team_stats_sheet = workbook.worksheets[0]

    # Write and calculate the total sheet
    total_sheet = workbook.copy_worksheet(team_stats_sheet)
    total_sheet.title = "TOTAL STATS"
    cell_values = calculate_numbers_for_cells(all_tags)
    for cell, value in cell_values.items():
        total_sheet[cell] = int(value)

    # Write sheets for individual games
    for _, tags_list in games_with_stats_dict.items():
        game_object = tags_list[0].game
        game_sheet = workbook.copy_worksheet(team_stats_sheet)
        game_sheet.title = f"{game_object.opponent} {game_object.date}"
        cell_values_for_game = calculate_numbers_for_cells(tags_list)
        for cell, value in cell_values_for_game.items():
            game_sheet[cell] = int(value)

    # Delete template sheet
    workbook.remove(workbook.worksheets[0])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=joukkuetilastot.xlsx"}
    )
########################## FOR TEAM STATS ##########################


########################## FOR PLAYER +/- ##########################
def build_game_data_structure(game: Game):
    data_structure = {
        "game_id": game.id,
        "opponent": game.opponent,
        "home": game.home,
        "date": game.date,
        "roster": {}}
    game_roster = data_structure["roster"]

    for roster_spot in game.in_rosters:
        player = roster_spot.player
        player_dict = {
            "name": f"{player.last_name.upper()} {player.first_name}",
            "position": player.position.name,
            "stats": {
                "ES-PG+": 0,   # ES-PG+ = Even strengthParticipated Goal +
                "ES-PG-": 0,
                "ES-PC+": 0,   
                "ES-PC-": 0,   # PC- = Partcipated chance -
                "ES-OIG+": 0,  
                "ES-OIG-": 0,  # OIG- = On ice Goal -
                "ES-OIC+": 0, 
                "ES-OIC-": 0,
                "PP-PG+": 0,   # PP-PG+ = Powerplay
                "PP-PG-": 0,
                "PP-PC+": 0,   
                "PP-PC-": 0,   
                "PP-OIG+": 0,  
                "PP-OIG-": 0,  
                "PP-OIC+": 0, 
                "PP-OIC-": 0,
                "PK-PG+": 0,   # PK-PG+ = Penaltykill
                "PK-PG-": 0,
                "PK-PC+": 0,   
                "PK-PC-": 0,
                "PK-OIG+": 0,  
                "PK-OIG-": 0,  
                "PK-OIC+": 0, 
                "PK-OIC-": 0
            }
        }
        game_roster[player.id] = player_dict

    return data_structure

def convert_roster_to_lists(roster: dict) -> dict:
    listed_roster = {
        "forwards": [],
        "defenders": []
    }

    for player_dict in roster.values():
        if player_dict["position"] == "FORWARD":
            listed_roster["forwards"].append(player_dict)
        elif player_dict["position"] == "DEFENDER":
            listed_roster["defenders"].append(player_dict)

    for group in listed_roster.values():
        group.sort(key=lambda player: player["name"])

    return listed_roster

def get_plusminus_games_data(teams_games: list[Game], db_session: Session):
    """
    Aggregates and processes game and player statistics data for a list of games.
    For each game in the provided list, this function:
    - Collects player statistics tags, on-ice links, and participating links from the database.
    - Maps and aggregates statistics for each player, both per-game and as a total across all games.
    - Handles statistics for different strengths (ES, PP, PK) and result types (goal for/against, chance for/against).
    - Converts roster dictionaries to lists for output compatibility.
    - Sorts the collected game data by date in descending order.
    Args:
        teams_games (list[Game]): List of Game objects to process.
        db_session (Session): SQLAlchemy database session for querying related data.
    Returns:
        tuple:
            - data_collector (list[dict]): List of processed game data dictionaries, one per game.
            - listed_total_data (list[dict]): Aggregated player statistics across all games as a list of dictionaries.
    """

    data_collector = []
    total_stats = {}
    game_ids = [game.id for game in teams_games]

    player_stats_tags = db_session.query(PlayerStatsTag).filter(PlayerStatsTag.game_id.in_(game_ids)).all()
    PSTs_by_game_id = defaultdict(list)
    for tag in player_stats_tags:
        PSTs_by_game_id[tag.game_id].append(tag)

    on_ice_links = db_session.query(PlayerStatsTagOnIce).join(PlayerStatsTag).filter(PlayerStatsTag.game_id.in_(game_ids)).all()  
    OILs_by_PST_id = defaultdict(list) # list of OnIce links with PlayerStatsTag id as key
    for on_ice_link in on_ice_links:
        OILs_by_PST_id[on_ice_link.tag_id].append(on_ice_link)

    participating_links = db_session.query(PlayerStatsTagParticipating).join(PlayerStatsTag).filter(PlayerStatsTag.game_id.in_(game_ids)).all()
    PLs_by_PST_id = defaultdict(list)
    for partic_link in participating_links:
        PLs_by_PST_id[partic_link.tag_id].append(partic_link)

    tag_type_mapping = {
                ShotResultTypes.GOAL_FOR: "G+",
                ShotResultTypes.GOAL_AGAINST: "G-",
                ShotResultTypes.CHANCE_FOR: "C+",
                ShotResultTypes.CHANCE_AGAINST: "C-",
                }



    for game in teams_games:
        game_data = build_game_data_structure(game)
        for id, data in game_data["roster"].items():
            if id not in total_stats:
                total_data = copy.deepcopy(data)
                total_data["GP"] = 1
                total_stats[id] = total_data
            else:
                total_stats[id]["GP"] += 1


        tags = PSTs_by_game_id[game.id]
        for tag in tags:
            strengths = tag.strengths
            if strengths not in ["ES", "PP", "PK"]:
                continue

            result_part = tag_type_mapping[tag.shot_result.value]
            
            tags_on_ice = OILs_by_PST_id[tag.id]
            OI_player_ids = [tag.player.id for tag in tags_on_ice]
            for id in OI_player_ids:
                game_data["roster"][id]["stats"][f"{strengths}-OI{result_part}"] += 1
                total_stats[id]["stats"][f"{strengths}-OI{result_part}"] += 1

            tags_participating = PLs_by_PST_id[tag.id]
            P_player_ids = [tag.player.id for tag in tags_participating]
            for id in P_player_ids:
                game_data["roster"][id]["stats"][f"{strengths}-P{result_part}"] += 1
                total_stats[id]["stats"][f"{strengths}-P{result_part}"] += 1


        game_data["roster"] = convert_roster_to_lists(game_data["roster"])

        data_collector.append(game_data)
    
    type(data_collector[0]["date"])
    data_collector.sort(key=lambda game: game["date"], reverse=True)
    listed_total_data = convert_roster_to_lists(total_stats)
    return data_collector, listed_total_data

@router.get("/plusminus")
async def get_plusminus_excel(game_ids: str = None, db_session: Session = Depends(get_db_session), current_user_id: int = Depends(get_current_user_id)):

    user = db_session.query(User).filter(User.id == current_user_id).first()
    team = user.team

    teams_games = team.games
    if game_ids:
        split_ids = [int(game_id) for game_id in game_ids.split(",")]
        teams_games = []
        for game in team.games:
            if game.id in split_ids:
                teams_games.append(game)
    else:
        teams_games = team.games
    

    data_for_games, total_stats = get_plusminus_games_data(teams_games, db_session)

    # Load the excel file and template sheet
    workbook = load_workbook("excels/plus_minus_template.xlsx")
    template_sheet = workbook.worksheets[0]

    stat_column_mapping = {
            "OIG+": "A",  
            "OIG-": "B",  # OIG- = On ice Goal -
            "OIC+": "D", 
            "OIC-": "E",
            "PG+": "H",   # PG+ = Participated Goal +
            "PG-": "I",
            "PC+": "K",   
            "PC-": "L",   # PC- = Partcipated chance -


            "ES-OIG+": "A",  
            "ES-OIG-": "B",  # OIG- = On ice Goal -
            "ES-OIC+": "D", 
            "ES-OIC-": "E",
            "ES-PG+": "H",   # ES-PG+ = Even strengthParticipated Goal +
            "ES-PG-": "I",
            "ES-PC+": "K",   
            "ES-PC-": "L",   # PC- = Partcipated chance -

            "PP-OIG+": "O",  
            "PP-OIG-": "P",  
            "PP-OIC+": "R", 
            "PP-OIC-": "S",
             "PP-PG+": "V",   # PP-PG+ = Powerplay
            "PP-PG-": "W",
            "PP-PC+": "Y",   
            "PP-PC-": "Z",   
   
            "PK-OIG+": "AC",  
            "PK-OIG-": "AD",  
            "PK-OIC+": "AF", 
            "PK-OIC-": "AG",
            "PK-PG+": "AJ",   # PK-PG+ = 
            "PK-PG-": "AK",
            "PK-PC+": "AM",   
            "PK-PC-": "AN"
    }
    name_colums = ["G", "U", "AI","AW"]

    total_sheet = workbook.copy_worksheet(template_sheet)
    total_sheet.title = "TOTAL"
    for i, defender in enumerate(total_stats["defenders"]):
        row = 4 + i
        for name_col in name_colums:
            total_sheet[f"{name_col}{row}"] = defender["name"]
        for key, value in defender["stats"].items():
            column = stat_column_mapping[key]
            total_sheet[f"{column}{row}"] = value

    for i, forward in enumerate(total_stats["forwards"]):
        row = 18 + i
        for name_col in name_colums:
            total_sheet[f"{name_col}{row}"] = forward["name"]
        for key, value in forward["stats"].items():
            column = stat_column_mapping[key]
            total_sheet[f"{column}{row}"] = value


    total_avg_sheet = workbook.copy_worksheet(template_sheet)
    total_avg_sheet.title = "TOTAL AVG"
    for i, defender in enumerate(total_stats["defenders"]):
        row = 4 + i
        games_played = defender["GP"]
        for name_col in name_colums:
            total_avg_sheet[f"{name_col}{row}"] = defender["name"]
        for key, value in defender["stats"].items():
            column = stat_column_mapping[key]
            total_avg_sheet[f"{column}{row}"] = value/games_played

    for i, forward in enumerate(total_stats["forwards"]):
        games_played = forward["GP"]
        row = 18 + i
        for name_col in name_colums:
            total_avg_sheet[f"{name_col}{row}"] = forward["name"]
        for key, value in forward["stats"].items():
            column = stat_column_mapping[key]
            total_avg_sheet[f"{column}{row}"] = value/games_played



    for game in data_for_games:
        game_sheet = workbook.copy_worksheet(template_sheet)
        game_sheet.title = f"{game["opponent"]} {game["date"]}"

        for i, defender in enumerate(game["roster"]["defenders"]):
            row = 4 + i
            for name_col in name_colums:
                game_sheet[f"{name_col}{row}"] = defender["name"]
            for key, value in defender["stats"].items():
                column = stat_column_mapping[key]
                game_sheet[f"{column}{row}"] = value

        for i, forward in enumerate(game["roster"]["forwards"]):
            row = 18 + i
            for name_col in name_colums:
                game_sheet[f"{name_col}{row}"] = forward["name"]
            for key, value in forward["stats"].items():
                column = stat_column_mapping[key]
                game_sheet[f"{column}{row}"] = value

            
    # Delete template sheet
    workbook.remove(workbook.worksheets[0])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plusminus.xlsx"}
    )

########################## FOR PLAYER +/- ##########################

########################## FOR TEAM SCORING  #######################
def get_outcome_cell_adjustment(tag: PlayerStatsTag) -> dict:
    """
    Calculates the adjustment to be made to a cell's row and column based on the outcome of a player's shot.
    Args:
        tag (PlayerStatsTag): An object containing information about the player's shot, including its result type.
    Returns:
        dict: A dictionary with 'row' and 'column' keys indicating the adjustment values to be applied.
    The adjustment is determined as follows:
        - If the shot result is CHANCE_AGAINST: column is incremented by 1.
        - If the shot result is GOAL_FOR: row is incremented by 1.
        - If the shot result is GOAL_AGAINST: both row and column are incremented by 1.
        - For CHANCE_FOR: no adjustment is made (row and column remain 0).
    """

    adjustment = {
        "row": 0,
        "column": 0
    }

    result_type = tag.shot_result.value

    if result_type == ShotResultTypes.CHANCE_AGAINST:
        adjustment["column"] = 1
    elif result_type == ShotResultTypes.GOAL_FOR:
        adjustment["row"] = 1
    elif result_type == ShotResultTypes.GOAL_AGAINST:
        adjustment["column"] = 1
        adjustment["row"] = 1

    return adjustment

def collect_shotzone_data(player_stats_tags: list[PlayerStatsTag]) -> dict:
    """
    Collects and aggregates shot zone data from a list of PlayerStatsTag objects.
    This function maps each shot area type to a specific Excel column, applies any necessary cell adjustments
    based on the outcome, and counts the occurrences of each resulting cell. The result is a dictionary
    mapping Excel cell references (e.g., "B5", "D6") to the number of times shots occurred in those zones.
    Args:
        player_stats_tags (list[PlayerStatsTag]): A list of PlayerStatsTag objects representing shot events.
    Returns:
        dict: A dictionary where keys are Excel cell references (str) and values are the counts (int) of shots in those cells.
    """

    ZONE_COLUMN_MAPPING = {
        ShotAreaTypes.ZONE_1: "B",
        ShotAreaTypes.ZONE_2_MIDDLE: "D",
        ShotAreaTypes.ZONE_2_SIDE: "F",
        ShotAreaTypes.HIGH_SLOT: "H",
        ShotAreaTypes.BLUELINE: "J",
        ShotAreaTypes.ZONE_4: "L",
        ShotAreaTypes.OUTSIDE_FAR: "N",
        ShotAreaTypes.OUTSIDE_CLOSE: "P",
        ShotAreaTypes.MISC: "R"
    }

    BASE_ROW = 5

    zone_cell_stats_dict = defaultdict(int)
    for tag in player_stats_tags:
        adjustment = get_outcome_cell_adjustment(tag)
        cell_col = chr(ord(ZONE_COLUMN_MAPPING[tag.shot_area.value]) + adjustment["column"])
        cell_row = BASE_ROW + adjustment["row"]
        cell = f"{cell_col}{cell_row}"
        zone_cell_stats_dict[cell] += 1
        
        # If its a goal, also increment the corresponding chance
        if cell_row == 6:
            zone_cell_stats_dict[f"{cell_col}{cell_row - 1}"] += 1

    return zone_cell_stats_dict

def collect_shot_type_data(player_stats_tags: list[PlayerStatsTag]) -> dict:
    """
    Collects and aggregates shot type data from a list of PlayerStatsTag objects, mapping each shot type to a specific Excel cell.
    Args:
        player_stats_tags (list[PlayerStatsTag]): A list of PlayerStatsTag objects representing individual shot events.
    Returns:
        dict: A dictionary where keys are Excel cell references (e.g., "B12") and values are the counts of shots mapped to those cells.
    Notes:
        - The mapping of shot types to Excel columns is defined in TYPE_COLUMN_MAPPING.
        - The row is determined by BASE_ROW and may be adjusted based on the outcome and whether the shot was cross-ice.
        - The function uses get_outcome_cell_adjustment to determine cell adjustments.
    """

    TYPE_COLUMN_MAPPING = {
        ShotTypeTypes.CARRY_SHOT: "B",
        ShotTypeTypes.CAN_SHOT: "D",
        ShotTypeTypes.ONE_TIMER: "F",
        ShotTypeTypes.LOWHIGH_SHOT: "N",
        ShotTypeTypes.TAKEAWAY_SHOT: "P",
        ShotTypeTypes.REBOUND_SHOT: "R",
        ShotTypeTypes.DEFLECTION_SHOT: "T"
    }

    BASE_ROW = 12
    type_cell_stats_dict = defaultdict(int)
    for tag in player_stats_tags:
        adjustment = get_outcome_cell_adjustment(tag)
        if tag.crossice:
            adjustment["column"] += 6
        cell_col = chr(ord(TYPE_COLUMN_MAPPING[tag.shot_type.value]) + adjustment["column"])
        cell_row = BASE_ROW + adjustment["row"]
        cell = f"{cell_col}{cell_row}"
        type_cell_stats_dict[cell] += 1

        # If its a goal, also increment the corresponding chance
        if cell_row == 13:
            type_cell_stats_dict[f"{cell_col}{cell_row - 1}"] += 1

    return type_cell_stats_dict

def collect_net_zone_data(player_stats_tags: list[PlayerStatsTag]) -> dict:
    WIDTH_COLUMN_MAPPING = {
        "Left": "C",
        "Mid": "D",
        "Right": "E",
    }

    HEIGHT_ROW_MAPPING = {
        "Top": 0,
        "Mid": 1,
        "Bottom": 2,
    }

    def get_adjustment(tag):
        adjustment = {
        "row": 0,
        "column": 0
    }
        result_type = tag.shot_result.value

        if result_type == ShotResultTypes.CHANCE_FOR:
            adjustment["column"] = 5
        elif result_type == ShotResultTypes.GOAL_AGAINST:
            adjustment["row"] = 7
        elif result_type == ShotResultTypes.CHANCE_AGAINST:
            adjustment["column"] = 5
            adjustment["row"] = 7

        return adjustment

    BASE_ROW = 19

    netzone_cell_stats_dict = defaultdict(int)
    for tag in player_stats_tags:
        adjustment = get_adjustment(tag)
        cell_col = chr(ord(WIDTH_COLUMN_MAPPING[tag.net_width]) + adjustment["column"])
        cell_row = BASE_ROW + HEIGHT_ROW_MAPPING[tag.net_height] + adjustment["row"]
        cell = f"{cell_col}{cell_row}"
        netzone_cell_stats_dict[cell] += 1

        if cell_col in ["C", "D", "E"]:
            chance_col = chr(ord(cell_col) + 5)
            netzone_cell_stats_dict[f"{chance_col}{cell_row}"] += 1

    return netzone_cell_stats_dict

def collect_shot_strengths_data(player_stats_tags: list[PlayerStatsTag]) -> dict:
    """
    Collects and aggregates shot strengths data from a list of PlayerStatsTag objects.
    This function maps each the number of players on the ice to a specific Excel cell based on predefined
    column mappings and calculated row/column adjustments. It then counts the occurrences of each
    cell reference, effectively summarizing the shot strengths distribution for later use in Excel export.
    Args:
        player_stats_tags (list[PlayerStatsTag]): A list of PlayerStatsTag objects containing shot strength information.
    Returns:
        dict: A dictionary where keys are Excel cell references (e.g., "B35") and values are the counts
              of shots mapped to each cell.
    """

    STRENGTHS_COLUMN_MAPPING = {
        "ES": "B",
        "PP": "D",
        "PK": "F",
        "EN+": "H",
        "EN-": "J",
    }

    BASE_ROW = 35
    strengths_cell_stats_dict = defaultdict(int)
    for tag in player_stats_tags:
        adjustment = get_outcome_cell_adjustment(tag)
        cell_col = chr(ord(STRENGTHS_COLUMN_MAPPING[tag.strengths]) + adjustment["column"])
        cell_row = BASE_ROW + adjustment["row"]
        cell = f"{cell_col}{cell_row}"
        strengths_cell_stats_dict[cell] += 1

        # If its a goal, also increment the corresponding chance
        if cell_row == 36:
            strengths_cell_stats_dict[f"{cell_col}{cell_row - 1}"] += 1

    return strengths_cell_stats_dict

def get_scoring_games_data(teams_games: list[Game], db_session: Session):
    """
    Collects and aggregates scoring-related statistics for a list of games.
    For each game in the provided list, this function gathers various player statistics
    (such as shot zones, shot types, net zones, and shot strengths) from the database,
    aggregates them per game, and also computes the totals across all games.
    Args:
        teams_games (list[Game]): A list of Game objects for which to collect statistics.
        db_session (Session): An active SQLAlchemy database session.
    Returns:
        tuple:
            - data_collector (list[dict]): A list of dictionaries, each containing:
                - "date": The date of the game.
                - "opponent": The opponent team.
                - "home": Boolean indicating if the game was at home.
                - "cell_values": A dictionary with aggregated statistics for the game.
            - total_cell_values (dict): A dictionary with aggregated statistics across all games.
    """


    game_ids = [game.id for game in teams_games]

    player_stats_tags = db_session.query(PlayerStatsTag).filter(PlayerStatsTag.game_id.in_(game_ids)).all()
    total_shot_zone_data = collect_shotzone_data(player_stats_tags)
    total_shot_type_data = collect_shot_type_data(player_stats_tags)
    total_net_zone_data = collect_net_zone_data(player_stats_tags)
    total_strengths_data = collect_shot_strengths_data(player_stats_tags)
    total_cell_values = {**total_shot_zone_data, **total_shot_type_data, **total_net_zone_data, **total_strengths_data}

    game_tags = defaultdict(list)
    for tag in player_stats_tags:
        game_tags[tag.game_id].append(tag)
    
    for game_id, tags in game_tags.items():
        print(f"{game_id}: has {len(tags)}")

    data_collector = []
    for game in teams_games:
        game_data = {}
        game_data["date"] = game.date
        game_data["opponent"] = game.opponent
        game_data["home"] = game.home

        tags_for_game = game_tags[game.id]
        shot_zone_data = collect_shotzone_data(tags_for_game)
        shot_type_data = collect_shot_type_data(tags_for_game)
        net_zone_data = collect_net_zone_data(tags_for_game)
        strengths_data = collect_shot_strengths_data(tags_for_game)
        
        game_cell_values = { **shot_zone_data, **shot_type_data, **net_zone_data, **strengths_data}
        game_data["cell_values"] = game_cell_values
        data_collector.append(game_data)

    return data_collector, total_cell_values

@router.get("/game-stats")
async def get_team_scoring_excel(game_ids: str = None, db_session: Session = Depends(get_db_session), current_user_id: int = Depends(get_current_user_id)):
    user = db_session.query(User).filter(User.id == current_user_id).first()
    team = user.team

    teams_games = team.games
    if game_ids:
        split_ids = [int(game_id) for game_id in game_ids.split(",")]
        teams_games = []
        for game in team.games:
            if game.id in split_ids:
                teams_games.append(game)
    else:
        teams_games = team.games
    

    cell_values_for_games, total_cell_values = get_scoring_games_data(teams_games, db_session)

    # Load the excel file and template sheet
    workbook = load_workbook("excels/game_stats_template.xlsx")
    template_sheet = workbook.worksheets[0]

    total_sheet = workbook.worksheets[1]
    for cell, value in total_cell_values.items():
        total_sheet[cell] = value

    cell_values_for_games.sort(key=lambda game: game["date"], reverse=True)
    for game in cell_values_for_games:
        game_sheet = workbook.copy_worksheet(template_sheet)
        game_sheet.title = f"{game["opponent"]} {game["date"]}"
        game_sheet["C1"] = game["date"]
        game_sheet["G1"] = game["opponent"]
        game_sheet["T1"] = game["home"]

        for cell, value in game["cell_values"].items():
            game_sheet[cell] = value
        
    # Delete template sheet
    workbook.remove(workbook.worksheets[0])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pelitilastot.xlsx"}
    )

########################## FOR TEAM SCORING  #######################


########################## FOR PLAYER SCORING  #####################
def collect_shooter_zones(player_stats_tags: list[PlayerStatsTag]) -> dict:
    ZONE_COLUMN_MAPPING = {
        ShotAreaTypes.ZONE_1: "B",
        ShotAreaTypes.ZONE_2_MIDDLE: "D",
        ShotAreaTypes.ZONE_2_SIDE: "F",
        ShotAreaTypes.HIGH_SLOT: "H",
        ShotAreaTypes.BLUELINE: "J",
        ShotAreaTypes.ZONE_4: "L",
        ShotAreaTypes.OUTSIDE_FAR: "N",
        ShotAreaTypes.OUTSIDE_CLOSE: "P",
        ShotAreaTypes.MISC: "R"
    }

    BASE_ROW = 4

    zone_cell_stats_dict = defaultdict(int)
    for tag in player_stats_tags:
        cell_col = ZONE_COLUMN_MAPPING[tag.shot_area.value]
        cell = f"{cell_col}{BASE_ROW}"
        zone_cell_stats_dict[cell] += 1
        
        # If its a goal
        if tag.shot_result.value == ShotResultTypes.GOAL_FOR:
            cell_col = chr(ord(ZONE_COLUMN_MAPPING[tag.shot_area.value]) + 1)
            zone_cell_stats_dict[f"{cell_col}{BASE_ROW}"] += 1

    return zone_cell_stats_dict

def collect_shooter_shot_types(player_stats_tags: list[PlayerStatsTag]) -> dict:
    TYPE_COLUMN_MAPPING = {
        ShotTypeTypes.CARRY_SHOT: "B",
        ShotTypeTypes.CAN_SHOT: "D",
        ShotTypeTypes.ONE_TIMER: "F",
        ShotTypeTypes.LOWHIGH_SHOT: "N",
        ShotTypeTypes.TAKEAWAY_SHOT: "P",
        ShotTypeTypes.REBOUND_SHOT: "R",
        ShotTypeTypes.DEFLECTION_SHOT: "T"
    }

    BASE_ROW = 11

    type_cell_stats_dict = defaultdict(int)
    for tag in player_stats_tags:
        cell_col = TYPE_COLUMN_MAPPING[tag.shot_type.value]
        cell = f"{cell_col}{BASE_ROW}"
        type_cell_stats_dict[cell] += 1
        
        # If its a goal
        if tag.shot_result.value == ShotResultTypes.GOAL_FOR:
            cell_col = chr(ord(TYPE_COLUMN_MAPPING[tag.shot_type.value]) + 1)
            type_cell_stats_dict[f"{cell_col}{BASE_ROW}"] += 1

    return type_cell_stats_dict

def collect_shooter_net_zones(player_stats_tags: list[PlayerStatsTag]) -> dict:
    WIDTH_COLUMN_MAPPING = {
        "Left": "G",
        "Mid": "H",
        "Right": "I",
    }

    HEIGHT_ROW_MAPPING = {
        "Top": 0,
        "Mid": 1,
        "Bottom": 2,
    }


    BASE_ROW = 17

    netzone_cell_stats_dict = defaultdict(int)
    for tag in player_stats_tags:
        cell_col = WIDTH_COLUMN_MAPPING[tag.net_width]
        cell_row = BASE_ROW + HEIGHT_ROW_MAPPING[tag.net_height]
        cell = f"{cell_col}{cell_row}"
        netzone_cell_stats_dict[cell] += 1

        if tag.shot_result.value == ShotResultTypes.GOAL_FOR:
            goal_col = chr(ord(cell_col) - 5)
            netzone_cell_stats_dict[f"{goal_col}{cell_row}"] += 1

    return netzone_cell_stats_dict

def collect_shooter_strengths(player_stats_tags: list[PlayerStatsTag]) -> dict:
    STRENGTHS_COLUMN_MAPPING = {
        "ES": "B",
        "PP": "C",
        "PK": "D",
        "EN+": "E",
        "EN-": "F",
    }

    BASE_ROW = 31
    strengths_cell_stats_dict = defaultdict(int)
    for tag in player_stats_tags:
        cell_col = STRENGTHS_COLUMN_MAPPING[tag.strengths]
        cell = f"{cell_col}{BASE_ROW}"
        strengths_cell_stats_dict[cell] += 1

        # If its a goal, also increment the corresponding chance
        if tag.shot_result.value == ShotResultTypes.GOAL_FOR:
            strengths_cell_stats_dict[f"{cell_col}{BASE_ROW + 1}"] += 1

    return strengths_cell_stats_dict

def find_players_on_ice_tags(player_stats_tags: list[PlayerStatsTag]) -> dict:
    """
    Aggregates a dictionary mapping each player ID to a list of PlayerStatsTag objects in which the player was on ice.
    Args:
        player_stats_tags (list[PlayerStatsTag]): A list of PlayerStatsTag objects, each containing information about players on ice.
    Returns:
        dict: A dictionary mapping each player ID to a list of PlayerStatsTag objects in which the player was on ice.
    """

    tags_for_each_player = defaultdict(list)

    for tag in player_stats_tags:
        for on_ice_tag in tag.players_on_ice:
            tags_for_each_player[on_ice_tag.player_id].append(tag)

    return tags_for_each_player

def collect_shooter_total_strengths(players_on_ice_tags: list[PlayerStatsTag], player_id: int ) -> dict:
    STRENGTHS_COLUMN_MAPPING = {
        "ES": "B",
        "PP": "E",
        "PK": "H",
        "EN+": "K",
        "EN-": "N",
    }


    strengths_cell_stats_dict = defaultdict(int)
    for tag in players_on_ice_tags:
        participating_player_ids = [participating_tag.player_id for participating_tag in tag.players_participating]

        if player_id in participating_player_ids:
            base_rows = [38, 46]
        else:
            base_rows = [46]

        adjustment = get_outcome_cell_adjustment(tag)
        for row in base_rows:
            cell_col = chr(ord(STRENGTHS_COLUMN_MAPPING[tag.strengths]) + adjustment["column"])
            cell_row = row + adjustment["row"]
            cell = f"{cell_col}{cell_row}"
            strengths_cell_stats_dict[cell] += 1

            # If its a goal, also increment the corresponding chance
            if tag.shot_result.value in [ShotResultTypes.GOAL_FOR, ShotResultTypes.GOAL_AGAINST]:
                strengths_cell_stats_dict[f"{cell_col}{cell_row - 1}"] += 1

    return strengths_cell_stats_dict

def collect_players_per_game_stats(players_stats_tags: list[PlayerStatsTag], player_id: int) -> dict:
    game_cell_values = {}
    for tag in players_stats_tags:
        if tag.game_id not in game_cell_values:
            game_cell_values[tag.game_id] = {}
            game_cell_values[tag.game_id]["date"] = tag.game.date
            game_cell_values[tag.game_id]["A"] = f"{tag.game.date} vs {tag.game.opponent}"
            for col in ["C", "D", "G", "H", "J", "K", "M", "N", "P", "Q", "T", "U", "W", "X", "Z", "AA", "AC", "AD", "AG", "AH", "AJ", "AK", "AM", "AN", "AP", "AQ"]:
                game_cell_values[tag.game_id][col] = 0

        if tag.shooter_id == player_id:
            game_cell_values[tag.game_id]["D"] += 1
            if tag.shot_result.value == ShotResultTypes.GOAL_FOR:
                game_cell_values[tag.game_id]["C"] += 1

        if tag.strengths == "ES":
            chance_col_ord = 71
        elif tag.strengths == "PP":
            chance_col_ord = 84
        elif tag.strengths == "PK":
            chance_col_ord = 97

        if tag.shot_result.value == ShotResultTypes.GOAL_AGAINST:
            chance_col_ord += 1
        elif tag.shot_result.value == ShotResultTypes.CHANCE_FOR:
            chance_col_ord += 3
        elif tag.shot_result.value == ShotResultTypes.CHANCE_AGAINST:
            chance_col_ord += 4

        if chance_col_ord > 90:
            chance_col = f"A{chr(chance_col_ord-26)}"
        else:
            chance_col = chr(chance_col_ord)
        game_cell_values[tag.game_id][chance_col] += 1
        
        participating_player_ids = [participating_tag.player_id for participating_tag in tag.players_participating]
        if player_id in participating_player_ids:
            participating_col_ord = chance_col_ord + 6
            if participating_col_ord > 90:
                participating_col = f"A{chr(participating_col_ord-26)}"
            else:
                participating_col = chr(participating_col_ord)
            game_cell_values[tag.game_id][participating_col] += 1

    sorted_games = sorted(game_cell_values.values(), key=lambda d: d["date"], reverse=True)
    return sorted_games

@router.get("/player-stats")
async def get_team_scoring_excel(game_ids: str = None, db_session: Session = Depends(get_db_session), current_user_id: int = Depends(get_current_user_id)):
    user = db_session.query(User).filter(User.id == current_user_id).first()
    team = user.team

    teams_games = team.games
    if game_ids:
        teams_games = []
        split_ids = [int(game_id) for game_id in game_ids.split(",")]
        for game in team.games:
            if game.id in split_ids:
                teams_games.append(game)
    selected_game_ids = [game.id for game in teams_games]

    player_stats_tags = db_session.query(PlayerStatsTag).filter(PlayerStatsTag.game_id.in_(selected_game_ids)).all()

    players_to_analyze = defaultdict(lambda: {"games": 0, "first_name": "", "last_name": "", "shooter_tags": [], "on_ice_tags": [], "cell_values": {}, "per_game_stats": []})
    for game in teams_games:

        for in_roster_object in game.in_rosters:
            player = in_roster_object.player
            if player.id not in players_to_analyze:
                players_to_analyze[player.id]["first_name"] = player.first_name 
                players_to_analyze[player.id]["last_name"] = player.last_name
            players_to_analyze[player.id]["games"] += 1
    
    for tag in player_stats_tags:
        if tag.shooter:
            players_to_analyze[tag.shooter.id]["shooter_tags"].append(tag)

    on_ice_tags_for_each_player = find_players_on_ice_tags(player_stats_tags) # {"player1_id": [PlayerStatsTags],..}
    for player_id, on_ice_tags in on_ice_tags_for_each_player.items():
        players_to_analyze[player_id]["on_ice_tags"] = on_ice_tags


    for player_id, player_data in players_to_analyze.items():
        zones = collect_shooter_zones(player_data["shooter_tags"])
        types = collect_shooter_shot_types(player_data["shooter_tags"])
        net_zones = collect_shooter_net_zones(player_data["shooter_tags"])
        strengths = collect_shooter_strengths(player_data["shooter_tags"])
        total_strengths = collect_shooter_total_strengths(player_data["on_ice_tags"], player_id)
        per_games_stats = collect_players_per_game_stats(player_data["on_ice_tags"], player_id)

        player_cell_values = {**zones, **types, **net_zones, **strengths, **total_strengths, "G1": player_data["games"]}
        players_to_analyze[player_id]["cell_values"] = player_cell_values
        players_to_analyze[player_id]["per_game_stats"] = per_games_stats


    # Load the excel file and template sheet
    workbook = load_workbook("excels/players_summary_template.xlsx")
    template_sheet = workbook.worksheets[0]

    for player_id, player_data in players_to_analyze.items():
        game_sheet = workbook.copy_worksheet(template_sheet)
        game_sheet.title = f"{player_data["last_name"].upper()} {player_data["first_name"]}"
        
        for cell, value in player_data["cell_values"].items():
            game_sheet[cell] = value
        
        for i, game in enumerate(player_data["per_game_stats"]):
            row = 54 + i
            for col, value in game.items():
                if col == "date":
                    continue
                cell = f"{col}{row}"
                game_sheet[cell] = value


    # Delete template sheet
    workbook.remove(workbook.worksheets[0])
    workbook.remove(workbook.worksheets[0])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pelaajayhteenveto.xlsx"}
    )
########################## FOR PLAYER SCORING  #####################
